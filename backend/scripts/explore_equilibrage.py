"""Explore les feuilles "EQUILIBRAGE" des fichiers Excel de BASE MANING.

Ce script ne modifie rien : il lit les fichiers en lecture seule et produit
un rapport pour savoir, avant d'ecrire le moindre import en base, quelles
colonnes existent vraiment et sous quels noms.

Usage:
    python explore_equilibrage.py
    python explore_equilibrage.py --limit 50
    python explore_equilibrage.py --csv rapport.csv
    python explore_equilibrage.py --data-dir "C:\\M1\\M2\\Soutenance\\Projet\\BASE MANING"
"""
import argparse
import csv
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

# Cle = nom canonique interne qu'on utilisera partout dans le code ensuite.
# Valeur = variantes deja observees (la casse/les accents/les espaces sont
# geres automatiquement par normalize_header, pas besoin de toutes les
# variantes de casse ici).
SYNONYMS = {
    "operateur": ["NOM OP", "NOM DES OP"],
    "operation": ["OPERATION"],
    # Choix valide avec l'utilisatrice : le temps de reference est le temps
    # APRES equilibrage de chaine, pas le temps chronometre brut.
    "temps": ["TEMPS EQUILIBREE"],
    "target": ["QTE/H/OPERATION", "TRGT CHRONO"],
    "machine": ["MATERIEL", "TYPE MACHNE", "MACHIE", "TYPE MACHINE"],
    # Numero de poste dans la chaine : sert potentiellement a ordonner les
    # operations de la gamme (a confirmer sur un exemple concret).
    "poste": ["POSTE N°"],
}

SHEET_NAME_TARGET = "EQUILIBRAGE"


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def normalize_header(raw) -> str:
    """Normalise un en-tete pour comparaison : majuscules, sans accents,
    espaces multiples ecrases, espaces autour des '/' retires."""
    if raw is None:
        return ""
    s = strip_accents(str(raw)).upper()
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\s*/\s*", "/", s)
    return s


_NORMALIZED_LOOKUP = {}
for canon, variants in SYNONYMS.items():
    for v in variants:
        _NORMALIZED_LOOKUP[normalize_header(v)] = canon


def to_canonical(raw_header):
    return _NORMALIZED_LOOKUP.get(normalize_header(raw_header))


def find_equilibrage_sheet(wb):
    """Retourne le nom de la feuille qui correspond a EQUILIBRAGE, peu
    importe la casse/les accents. Tolere aussi un nom qui la contient
    (ex: 'EQUILIBRAGE 2023')."""
    target = normalize_header(SHEET_NAME_TARGET)
    for name in wb.sheetnames:
        if normalize_header(name) == target:
            return name
    for name in wb.sheetnames:
        if target in normalize_header(name):
            return name
    return None


def find_header_row(ws, max_rows_to_scan=15):
    """Les fichiers de production ont parfois des lignes de titre avant la
    vraie ligne d'entete. On prend la ligne qui reconnait le plus de
    colonnes canoniques parmi les premieres lignes."""
    best_row_idx, best_score = None, 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows_to_scan), start=1):
        score = sum(1 for cell in row if to_canonical(cell.value))
        if score > best_score:
            best_score, best_row_idx = score, row_idx
    return best_row_idx, best_score


def find_best_sheet_by_content(wb, min_score=3):
    """Repli quand aucune feuille ne s'appelle EQUILIBRAGE : on regarde le
    CONTENU de toutes les feuilles du classeur et on garde celle dont la
    ligne d'entete reconnait le plus de colonnes canoniques. Les noms de
    feuille rencontres dans ce jeu de donnees sont trop varies (EQUI,
    "EQUILIBRE A 70%", nom du modele...) pour etre tous listes a la main."""
    best_name, best_row_idx, best_score = None, None, 0
    for name in wb.sheetnames:
        ws = wb[name]
        if not hasattr(ws, "iter_rows"):
            # Feuille graphique (Chartsheet) ou autre type sans cellules.
            continue
        row_idx, score = find_header_row(ws)
        if score > best_score:
            best_name, best_row_idx, best_score = name, row_idx, score
    if best_score >= min_score:
        return best_name, best_row_idx
    return None, None


def explore(data_dir: Path, limit=None):
    files = sorted(data_dir.rglob("*.xlsx")) + sorted(data_dir.rglob("*.xlsm"))
    if limit:
        files = files[:limit]

    report_rows = []
    unmatched_headers = Counter()
    canonical_found_count = Counter()
    files_with_sheet = 0
    match_method_count = Counter()
    files_without_sheet = []
    files_with_errors = []

    for path in files:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            files_with_errors.append((path, str(exc)))
            continue

        sheet_name = find_equilibrage_sheet(wb)
        if sheet_name is not None and hasattr(wb[sheet_name], "iter_rows"):
            match_method = "nom"
            header_row_idx, _score = find_header_row(wb[sheet_name])
        else:
            sheet_name, header_row_idx = find_best_sheet_by_content(wb)
            match_method = "contenu" if sheet_name else None

        if sheet_name is None:
            files_without_sheet.append(path)
            wb.close()
            continue

        files_with_sheet += 1
        match_method_count[match_method] += 1
        ws = wb[sheet_name]

        found_canonical = set()
        row_report = {
            "fichier": str(path.relative_to(data_dir)),
            "feuille": sheet_name,
            "trouvee_par": match_method,
            "ligne_entete": header_row_idx,
        }

        if header_row_idx:
            header_cells = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))
            for cell in header_cells:
                raw = cell.value
                if raw is None or str(raw).strip() == "":
                    continue
                canon = to_canonical(raw)
                if canon:
                    found_canonical.add(canon)
                else:
                    unmatched_headers[normalize_header(raw)] += 1

        # Une seule occurrence comptee par fichier, meme si la colonne
        # apparait deux fois dans la meme feuille.
        for canon in found_canonical:
            canonical_found_count[canon] += 1

        for canon in SYNONYMS:
            row_report[canon] = "OK" if canon in found_canonical else "MANQUANT"

        report_rows.append(row_report)
        wb.close()

    return {
        "report_rows": report_rows,
        "unmatched_headers": unmatched_headers,
        "canonical_found_count": canonical_found_count,
        "files_with_sheet": files_with_sheet,
        "match_method_count": match_method_count,
        "files_without_sheet": files_without_sheet,
        "files_with_errors": files_with_errors,
        "total_files": len(files),
    }


def print_summary(result):
    print(f"Fichiers analyses (.xlsx/.xlsm) : {result['total_files']}")
    print(f"Fichiers avec une feuille de gamme detectee : {result['files_with_sheet']}")
    print(f"  - trouvee par nom de feuille (EQUILIBRAGE) : {result['match_method_count'].get('nom', 0)}")
    print(f"  - trouvee par contenu (repli, nom different) : {result['match_method_count'].get('contenu', 0)}")
    print(f"Fichiers SANS feuille de gamme detectee : {len(result['files_without_sheet'])}")
    print(f"Fichiers en erreur (illisibles) : {len(result['files_with_errors'])}")
    print()
    print("Colonnes canoniques trouvees (sur combien de fichiers ayant la feuille) :")
    for canon in SYNONYMS:
        print(f"  - {canon:12s}: {result['canonical_found_count'].get(canon, 0)} fichiers")
    print()
    print("En-tetes NON reconnus (a ajouter aux synonymes si pertinent), tries par frequence :")
    for header, count in result["unmatched_headers"].most_common(40):
        print(f"  {count:4d}  {header}")
    if result["files_without_sheet"]:
        print()
        print(f"Exemples de fichiers sans feuille EQUILIBRAGE (10 sur {len(result['files_without_sheet'])}) :")
        for p in result["files_without_sheet"][:10]:
            print(f"  {p}")
    if result["files_with_errors"]:
        print()
        print(f"Fichiers en erreur (10 sur {len(result['files_with_errors'])}) :")
        for p, err in result["files_with_errors"][:10]:
            print(f"  {p} -> {err}")


def write_csv(result, out_path: Path):
    if not result["report_rows"]:
        return
    fieldnames = list(result["report_rows"][0].keys())
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(result["report_rows"])


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data-dir", type=Path, default=None, help='Chemin vers le dossier "BASE MANING"')
    parser.add_argument("--limit", type=int, default=None, help="Limiter le nombre de fichiers (test rapide)")
    parser.add_argument("--csv", type=Path, default=None, help="Chemin du CSV de sortie detaille (un fichier = une ligne)")
    args = parser.parse_args()

    if args.data_dir is None:
        # backend/scripts/ -> backend -> Gamme_montage_M2 -> Projet -> "BASE MANING"
        args.data_dir = Path(__file__).resolve().parents[3] / "BASE MANING"

    if not args.data_dir.exists():
        raise SystemExit(f"Dossier introuvable : {args.data_dir}")

    result = explore(args.data_dir, limit=args.limit)
    print_summary(result)

    if args.csv:
        write_csv(result, args.csv)
        print(f"\nRapport detaille ecrit dans : {args.csv}")


if __name__ == "__main__":
    main()
