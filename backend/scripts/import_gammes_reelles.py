"""Importe le CONTENU reel des gammes depuis BASE MANING dans PostgreSQL.

Reutilise la logique de detection de feuille/en-tetes deja validee dans
explore_equilibrage.py (meme dossier). Ce script-ci va plus loin : il lit
les lignes de donnees (pas juste les en-tetes) et les insere en base.

Regles de mapping (identiques a l'exploration) :
    - operateur : colonne "Nom op" / "Nom des op"
    - operation : colonne "OPERATION"
    - temps     : colonne "TEMPS EQUILIBREE" (temps APRES equilibrage de
                  chaine, choix valide - pas le temps chronometre brut)
    - target    : colonne "QTE/H/OPERATION" ou "TRGT CHRONO"
    - machine   : colonne "MATERIEL" / "TYPE MACHINE" / etc.
    - ordre     : position reelle de la ligne dans le tableau (le numero de
                  poste Excel n'est pas fiable, on ne s'en sert jamais)

Chaque fichier Excel = un Article + une Gamme (1 pour 1). Le code article
est derive du chemin relatif du fichier (garantit l'unicite meme quand
plusieurs fichiers portent le meme nom dans des dossiers differents).

Idempotent par defaut : un fichier deja importe (code deja present en
base) est saute. Utiliser --reset-reel pour repartir de zero (supprime
uniquement les articles reels, jamais les articles DEMO-*).

Usage:
    python import_gammes_reelles.py --limit 30          # test rapide
    python import_gammes_reelles.py                      # import complet
    python import_gammes_reelles.py --reset-reel          # reimport complet propre
"""
import argparse
import re
import sys
import time
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))       # backend/scripts (explore_equilibrage)
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))   # backend/ (database, models)

from explore_equilibrage import (  # noqa: E402
    find_best_sheet_by_content,
    find_equilibrage_sheet,
    find_header_row,
    to_canonical,
)
from database import Base_chebdo, SessionLocal, engine  # noqa: E402
from models import Article, Gamme, GammeLigne, Operateur  # noqa: E402


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def normalize_libelle(raw) -> str:
    return re.sub(r"\s+", " ", strip_accents(str(raw)).upper()).strip()


def build_code(rel_path: Path) -> str:
    """Code unique et stable derive du chemin relatif du fichier."""
    s = str(rel_path.with_suffix(""))
    s = re.sub(r"\s+", " ", s).strip()
    if len(s) > 150:
        s = s[:150]
    return s


# Une chaine a ses propres operateurs : on ne suggere jamais un operateur
# d'une autre chaine (meme numero de chaine reutilise par des dossiers
# racine differents = lignes physiquement differentes -> on combine les
# deux : "RACINE::CHn").
CHAIN_PATTERN = re.compile(r"(CHAINE|CH|LINE|L)\.?\s*0*(\d+)", re.IGNORECASE)


def extraire_chaine(rel_path: Path) -> str | None:
    segments = str(rel_path).split("\\")
    if len(segments) < 2:
        return None
    racine = segments[0].strip().upper()
    for segment in segments[1:-1]:
        seg = segment.strip()
        if re.fullmatch(r"\d{4}", seg):  # annee
            continue
        m = CHAIN_PATTERN.search(seg)
        if m:
            num = m.group(2).lstrip("0") or "0"
            return f"{racine}::CH{num}"
    return None


def parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s or s in ("-", "."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normaliser_nom_operateur(nom: str) -> str:
    """Cle de dedup : espaces ecrases, casse et accents ignores."""
    return re.sub(r"\s+", " ", strip_accents(nom).upper()).strip()


def detecter_gamme(wb):
    """Reutilise la meme logique que explore_equilibrage.py : nom de
    feuille EQUILIBRAGE en priorite, sinon repli par contenu."""
    sheet_name = find_equilibrage_sheet(wb)
    if sheet_name is not None and hasattr(wb[sheet_name], "iter_rows"):
        header_row_idx, _score = find_header_row(wb[sheet_name])
        return sheet_name, header_row_idx
    return find_best_sheet_by_content(wb)


def construire_index_colonnes(ws, header_row_idx):
    """Colonne (1-based) pour chaque champ canonique, a partir de la ligne
    d'entete detectee. En cas de doublon, on garde la premiere occurrence."""
    index = {}
    header_cells = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))
    for cell in header_cells:
        canon = to_canonical(cell.value)
        if canon and canon not in index:
            index[canon] = cell.column
    return index


# Restes de zone de signature ou de bloc "fiche de suivi" (chef de chaine,
# effectif, min produite/presence...) qui atterrissent parfois dans la
# colonne OPERATION en bas de tableau. Liste volontairement restreinte a
# des cas surs : on ne filtre jamais un libelle qui ressemble a une vraie
# operation (ex: REPASSAGE, BOUT DE CHAINE), meme incomplet par ailleurs.
LIBELLES_PARASITES = {
    "IE", "HH", "AA", "II", "PA", "EE", "GG",
    "PREPARE PAR", "VERIFIE PAR", "APPROUVE PAR", "CONTROLE PAR",
    "CHEF DE CHAINE", "VISER PAR :", "VISER PAR", "TECHNICAL",
    "GUIDE ET PIEDS", "NOMBRE", "NOMBRES", "MIN PRODUITE", "MIN PRESENCE",
    "QTE PAQUET",
}


def extraire_lignes(ws, header_row_idx, colonnes):
    """Genere les lignes de gamme valides (operation non vide), dans
    l'ordre reel du tableau."""
    col_operation = colonnes.get("operation")
    if col_operation is None:
        return

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        get = lambda canon: row[colonnes[canon] - 1].value if canon in colonnes else None  # noqa: E731

        operation_libelle = get("operation")
        if operation_libelle is None or str(operation_libelle).strip() == "":
            continue
        libelle_normalise = normalize_libelle(operation_libelle)
        if libelle_normalise in LIBELLES_PARASITES:
            continue
        if re.fullmatch(r"\d+", libelle_normalise):  # nombre isole (ex: "30")
            continue

        operateur_nom = get("operateur")
        operateur_nom = str(operateur_nom).strip() if operateur_nom is not None else None
        if operateur_nom == "":
            operateur_nom = None

        machine = get("machine")
        machine = str(machine).strip() if machine is not None else None
        if machine == "":
            machine = None

        yield {
            "operation_libelle": str(operation_libelle).strip(),
            "operateur_nom": operateur_nom,
            "temps_equilibre": parse_float(get("temps")),
            "target": parse_float(get("target")),
            "machine": machine,
        }


class CacheOperateurs:
    """Evite une requete DB par ligne : precharge les operateurs existants
    et cree les nouveaux au fil de l'import (un seul flush par nouveau)."""

    def __init__(self, db):
        self.db = db
        self._par_cle = {
            normaliser_nom_operateur(op.nom): op for op in db.query(Operateur).all()
        }

    def get_or_create(self, nom: str) -> Operateur:
        cle = normaliser_nom_operateur(nom)
        operateur = self._par_cle.get(cle)
        if operateur is None:
            operateur = Operateur(nom=nom, actif=True)
            self.db.add(operateur)
            self.db.flush()
            self._par_cle[cle] = operateur
        return operateur


def reset_donnees_reelles(db):
    articles = db.query(Article).filter(~Article.code.like("DEMO-%")).all()
    for article in articles:
        db.delete(article)  # cascade -> gamme -> gamme_ligne
    db.commit()


def importer(data_dir: Path, limit=None, commit_every=25):
    Base_chebdo.metadata.create_all(engine)
    db = SessionLocal()
    cache_operateurs = CacheOperateurs(db)
    codes_existants = {c for (c,) in db.query(Article.code).all()}

    files = sorted(data_dir.rglob("*.xlsx")) + sorted(data_dir.rglob("*.xlsm"))
    files = [p for p in files if not p.name.startswith("~$")]
    if limit:
        files = files[:limit]

    stats = {
        "traites": 0,
        "deja_importes": 0,
        "importes": 0,
        "sans_feuille": 0,
        "erreurs": 0,
        "lignes_creees": 0,
        "operateurs_crees_avant": len(cache_operateurs._par_cle),
    }

    debut = time.time()
    depuis_dernier_commit = 0

    for i, path in enumerate(files, start=1):
        stats["traites"] += 1
        rel_path = path.relative_to(data_dir)
        code = build_code(rel_path)

        if code in codes_existants:
            stats["deja_importes"] += 1
            continue

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            stats["erreurs"] += 1
            print(f"[ERREUR] {rel_path} -> {exc}")
            continue

        try:
            sheet_name, header_row_idx = detecter_gamme(wb)
            if sheet_name is None or header_row_idx is None:
                stats["sans_feuille"] += 1
                continue

            ws = wb[sheet_name]
            colonnes = construire_index_colonnes(ws, header_row_idx)
            lignes = list(extraire_lignes(ws, header_row_idx, colonnes))

            if not lignes:
                stats["sans_feuille"] += 1
                continue

            article = Article(
                code=code,
                nom=path.stem.strip(),
                description=None,
                source=str(rel_path),
                chaine=extraire_chaine(rel_path),
            )
            db.add(article)
            db.flush()

            gamme = Gamme(article_id=article.id)
            db.add(gamme)
            db.flush()

            for ordre, ligne in enumerate(lignes, start=1):
                operateur_id = None
                if ligne["operateur_nom"]:
                    operateur_id = cache_operateurs.get_or_create(ligne["operateur_nom"]).id

                db.add(
                    GammeLigne(
                        gamme_id=gamme.id,
                        ordre=ordre,
                        operation_libelle=ligne["operation_libelle"],
                        temps_equilibre=ligne["temps_equilibre"],
                        target=ligne["target"],
                        machine=ligne["machine"],
                        operateur_id=operateur_id,
                    )
                )

            stats["lignes_creees"] += len(lignes)
            stats["importes"] += 1
            codes_existants.add(code)
            depuis_dernier_commit += 1

        except Exception as exc:
            db.rollback()
            stats["erreurs"] += 1
            print(f"[ERREUR] {rel_path} -> {exc}")
            continue
        finally:
            wb.close()

        if depuis_dernier_commit >= commit_every:
            db.commit()
            depuis_dernier_commit = 0

        if i % 200 == 0:
            ecoule = time.time() - debut
            print(
                f"... {i}/{len(files)} fichiers traites "
                f"({stats['importes']} importes, {stats['lignes_creees']} lignes, "
                f"{ecoule:.0f}s ecoulees)"
            )

    db.commit()
    db.close()

    stats["duree_s"] = round(time.time() - debut, 1)
    return stats


def print_stats(stats):
    print()
    print("=== Import termine ===")
    print(f"Fichiers traites          : {stats['traites']}")
    print(f"Deja importes (ignores)   : {stats['deja_importes']}")
    print(f"Nouveaux articles importes: {stats['importes']}")
    print(f"Lignes de gamme creees    : {stats['lignes_creees']}")
    print(f"Sans gamme exploitable    : {stats['sans_feuille']}")
    print(f"Erreurs                   : {stats['erreurs']}")
    print(f"Duree                     : {stats['duree_s']}s")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--data-dir", type=Path, default=None)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--reset-reel", action="store_true", help="Supprime les articles reels existants avant de reimporter (jamais les DEMO-*)")
    args = parser.parse_args()

    if args.data_dir is None:
        args.data_dir = Path(__file__).resolve().parents[3] / "BASE MANING"
    if not args.data_dir.exists():
        raise SystemExit(f"Dossier introuvable : {args.data_dir}")

    if args.reset_reel:
        Base_chebdo.metadata.create_all(engine)
        db = SessionLocal()
        reset_donnees_reelles(db)
        db.close()
        print("Donnees reelles existantes supprimees (DEMO-* conservees).")

    stats = importer(args.data_dir, limit=args.limit)
    print_stats(stats)


if __name__ == "__main__":
    main()
