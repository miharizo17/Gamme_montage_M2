"""Generation de fichiers Excel/PDF telechargeables pour une gamme de
montage enregistree (ArticleDetailOut), utilisee par les boutons d'export
de l'historique. Ne touche pas a la base : pure mise en forme."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from schema import ArticleDetailOut

COULEUR_ACCENT = colors.HexColor("#9c6b52")
COULEUR_BORDURE = colors.HexColor("#e6d9c8")


def _smv_minutes(article: ArticleDetailOut) -> float:
    return sum(ligne.temps_equilibre or 0 for ligne in article.lignes_gamme) / 60


def generer_excel_gamme(article: ArticleDetailOut) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gamme de montage"

    ws["A1"] = article.nom
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Code : {article.code}"
    ws["A3"] = f"Chaine : {article.chaine or '-'}"
    ws["A4"] = f"Date d'enregistrement : {article.date_creation.strftime('%d/%m/%Y') if article.date_creation else '-'}"

    ligne_entete = 6
    for col, texte in enumerate(["#", "Operation", "Machine", "Temps (s)", "Operateur"], start=1):
        cellule = ws.cell(row=ligne_entete, column=col, value=texte)
        cellule.font = Font(bold=True)

    for i, ligne in enumerate(article.lignes_gamme, start=1):
        row = ligne_entete + i
        ws.cell(row=row, column=1, value=ligne.ordre)
        ws.cell(row=row, column=2, value=ligne.operation_libelle)
        ws.cell(row=row, column=3, value=ligne.machine or "-")
        ws.cell(row=row, column=4, value=ligne.temps_equilibre)
        ws.cell(row=row, column=5, value=ligne.operateur_nom or "-")

    ligne_total = ligne_entete + len(article.lignes_gamme) + 1
    ws.cell(row=ligne_total, column=2, value="SMV totale (min)").font = Font(bold=True)
    ws.cell(row=ligne_total, column=4, value=round(_smv_minutes(article), 1))

    for lettre, largeur in zip("ABCDE", (6, 42, 18, 12, 22)):
        ws.column_dimensions[lettre].width = largeur

    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


def generer_pdf_gamme(article: ArticleDetailOut) -> bytes:
    tampon = io.BytesIO()
    doc = SimpleDocTemplate(tampon, pagesize=A4, title=f"Gamme {article.code}")
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(article.nom, styles["Title"]),
        Paragraph(f"Code : {article.code} — Chaine : {article.chaine or '-'}", styles["Normal"]),
    ]
    if article.date_creation:
        elements.append(Paragraph(f"Date d'enregistrement : {article.date_creation.strftime('%d/%m/%Y')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    donnees = [["#", "Operation", "Machine", "Temps (s)", "Operateur"]]
    for ligne in article.lignes_gamme:
        donnees.append(
            [
                str(ligne.ordre),
                ligne.operation_libelle,
                ligne.machine or "-",
                f"{ligne.temps_equilibre:.1f}" if ligne.temps_equilibre is not None else "-",
                ligne.operateur_nom or "-",
            ]
        )

    table = Table(donnees, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), COULEUR_ACCENT),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, COULEUR_BORDURE),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fbf3ea")]),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(f"SMV totale : {_smv_minutes(article):.1f} min", styles["Normal"]))

    doc.build(elements)
    return tampon.getvalue()
